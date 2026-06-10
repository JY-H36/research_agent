"""
图存储模块
- MySQL 持久化存储（实体 + 关系 CRUD）
- NetworkX 内存图（查询缓存，读时重建）
"""
import json
from typing import Dict, List, Optional, Tuple
import networkx as nx

from database.connection import SessionLocal
from knowledge_graph.models import (
    KgPaper, KgAuthor, KgDataset, KgMethod, KgMetric, KgTask, KgVenue,
    KgPaperAuthor, KgPaperCites, KgPaperUsesMethod,
    KgPaperEvaluatesOn, KgPaperTrainsOn, KgPaperBelongsToTask,
    KgMethodImprovesMethod, KgPaperPublishedIn, KgPaperReportsMetric,
    PaperSource,
)
from utils.logger import get_logger

logger = get_logger(__name__)

# ============================================================
# NetworkX 图缓存
# ============================================================
_nx_graph: Optional[nx.MultiDiGraph] = None
_graph_dirty: bool = True


def _mark_graph_dirty():
    global _graph_dirty
    _graph_dirty = True


def get_nx_graph() -> nx.MultiDiGraph:
    """获取 NetworkX 图，如果脏则重建"""
    global _nx_graph, _graph_dirty
    if _nx_graph is None or _graph_dirty:
        _nx_graph = _build_nx_graph()
        _graph_dirty = False
        logger.info("NetworkX 图构建完成: %d 节点, %d 边",
                    _nx_graph.number_of_nodes(), _nx_graph.number_of_edges())
    return _nx_graph


def _build_nx_graph() -> nx.MultiDiGraph:
    """从 MySQL 加载所有实体和关系，构建 NetworkX MultiDiGraph"""
    G = nx.MultiDiGraph()
    db = SessionLocal()
    try:
        # --- 节点: Papers ---
        for p in db.query(KgPaper).all():
            G.add_node(f"paper:{p.paper_id}", type="paper", label=p.title[:200],
                       year=p.year, source=p.source.value if p.source else "upload",
                       citation_count=p.citation_count or 0)

        # --- 节点: Authors ---
        for a in db.query(KgAuthor).all():
            G.add_node(f"author:{a.author_id}", type="author", label=a.name,
                       affiliation=a.affiliation)

        # --- 节点: Methods ---
        for m in db.query(KgMethod).all():
            aliases = m.aliases or []
            if isinstance(aliases, str):
                aliases = json.loads(aliases)
            G.add_node(f"method:{m.method_id}", type="method", label=m.name,
                       category=m.category or "",
                       aliases=aliases)

        # --- 节点: Datasets ---
        for d in db.query(KgDataset).all():
            G.add_node(f"dataset:{d.dataset_id}", type="dataset", label=d.name,
                       domain=d.domain, task=d.task)

        # --- 节点: Tasks ---
        for t in db.query(KgTask).all():
            G.add_node(f"task:{t.task_id}", type="task", label=t.name,
                       level=t.level, parent_task_id=t.parent_task_id)

        # --- 节点: Metrics ---
        for m in db.query(KgMetric).all():
            G.add_node(f"metric:{m.metric_id}", type="metric", label=m.name,
                       direction=m.direction.value if m.direction else "")

        # --- 节点: Venues ---
        for v in db.query(KgVenue).all():
            G.add_node(f"venue:{v.venue_id}", type="venue", label=v.name,
                       abbreviation=v.abbreviation,
                       venue_type=v.type.value if v.type else "")

        # --- 边: WRITTEN_BY ---
        for r in db.query(KgPaperAuthor).all():
            G.add_edge(f"paper:{r.paper_id}", f"author:{r.author_id}",
                       relation="WRITTEN_BY", author_order=r.author_order,
                       is_corresponding=r.is_corresponding)

        # --- 边: CITES ---
        for r in db.query(KgPaperCites).all():
            G.add_edge(f"paper:{r.citing_paper_id}", f"paper:{r.cited_paper_id}",
                       relation="CITES", citation_type=r.citation_type.value if r.citation_type else "",
                       section=r.section)

        # --- 边: USES_METHOD ---
        for r in db.query(KgPaperUsesMethod).all():
            G.add_edge(f"paper:{r.paper_id}", f"method:{r.method_id}",
                       relation="USES_METHOD", role=r.role, variant=r.variant,
                       perf=r.performance_contribution.value if r.performance_contribution else "")

        # --- 边: EVALUATES_ON ---
        for r in db.query(KgPaperEvaluatesOn).all():
            metrics = r.metrics or {}
            if isinstance(metrics, str):
                metrics = json.loads(metrics)
            G.add_edge(f"paper:{r.paper_id}", f"dataset:{r.dataset_id}",
                       relation="EVALUATES_ON", task=r.task, split=r.split,
                       metrics=metrics, protocol=r.protocol)

        # --- 边: TRAINS_ON ---
        for r in db.query(KgPaperTrainsOn).all():
            G.add_edge(f"paper:{r.paper_id}", f"dataset:{r.dataset_id}",
                       relation="TRAINS_ON", split=r.split)

        # --- 边: BELONGS_TO ---
        for r in db.query(KgPaperBelongsToTask).all():
            G.add_edge(f"paper:{r.paper_id}", f"task:{r.task_id}",
                       relation="BELONGS_TO", is_primary=r.is_primary)

        # --- 边: IMPROVES_UPON ---
        for r in db.query(KgMethodImprovesMethod).all():
            G.add_edge(f"method:{r.method_a_id}", f"method:{r.method_b_id}",
                       relation="IMPROVES_UPON", description=r.description or "")

        # --- 边: PUBLISHED_IN ---
        for r in db.query(KgPaperPublishedIn).all():
            G.add_edge(f"paper:{r.paper_id}", f"venue:{r.venue_id}",
                       relation="PUBLISHED_IN", date=r.date)

        # --- 边: REPORTS_METRIC ---
        for r in db.query(KgPaperReportsMetric).all():
            G.add_edge(f"paper:{r.paper_id}", f"metric:{r.metric_id}",
                       relation="REPORTS_METRIC", value=r.value,
                       dataset_id=r.dataset_id, condition=r.condition)

    finally:
        db.close()
    return G


# ============================================================
# 实体 CRUD
# ============================================================

def add_paper(paper_data: Dict) -> str:
    """
    创建 Paper 节点。
    paper_data 可包含: title, abstract, abstract_cn, keywords, authors,
                       method_name, method_summary, year, doi, arxiv_id,
                       url, pdf_url, venue_name, citation_count,
                       language, source, document_id, docling_md
    返回 paper_id
    """
    db = SessionLocal()
    try:
        paper = KgPaper(
            title=paper_data.get("title", "Untitled"),
            abstract=paper_data.get("abstract", ""),
            abstract_cn=paper_data.get("abstract_cn", ""),
            keywords=paper_data.get("keywords", []),
            authors=paper_data.get("authors", []),
            method_name=paper_data.get("method_name", ""),
            method_summary=paper_data.get("method_summary", ""),
            year=paper_data.get("year"),
            doi=paper_data.get("doi", ""),
            arxiv_id=paper_data.get("arxiv_id", ""),
            url=paper_data.get("url", ""),
            pdf_url=paper_data.get("pdf_url", ""),
            venue_name=paper_data.get("venue_name", ""),
            citation_count=paper_data.get("citation_count", 0),
            language=paper_data.get("language", "en"),
            source=paper_data.get("source", PaperSource.UPLOAD),
            document_id=paper_data.get("document_id"),
            docling_md=paper_data.get("docling_md", ""),
        )
        db.add(paper)
        db.commit()
        paper_id = paper.paper_id
        logger.info("Paper 入图: %s", paper_data.get("title", "")[:100])
        _mark_graph_dirty()
        return paper_id
    except Exception as e:
        db.rollback()
        logger.error("Paper 入图失败: %s", e, exc_info=True)
        raise
    finally:
        db.close()


def add_author(name: str, affiliation: str = "", orcid: str = "") -> str:
    """创建或查找已有 Author（按 name + affiliation 去重）"""
    db = SessionLocal()
    try:
        existing = db.query(KgAuthor).filter(
            KgAuthor.name == name,
            KgAuthor.affiliation == affiliation,
        ).first()
        if existing:
            return existing.author_id

        author = KgAuthor(name=name, affiliation=affiliation, orcid=orcid)
        db.add(author)
        db.commit()
        _mark_graph_dirty()
        return author.author_id
    except Exception as e:
        db.rollback()
        logger.error("Author 入图失败: %s", e)
        raise
    finally:
        db.close()


def add_method(name: str, category: str = "", aliases: List[str] = None,
               description: str = "") -> str:
    """创建或查找已有 Method（按 name 去重）"""
    db = SessionLocal()
    try:
        existing = db.query(KgMethod).filter(KgMethod.name == name).first()
        if existing:
            # 合并 aliases
            if aliases:
                current_aliases = existing.aliases or []
                if isinstance(current_aliases, str):
                    current_aliases = json.loads(current_aliases)
                merged = list(set(current_aliases + aliases))
                existing.aliases = merged
                db.commit()
            return existing.method_id

        method = KgMethod(name=name, aliases=aliases or [], category=category or None,
                          description=description)
        db.add(method)
        db.commit()
        _mark_graph_dirty()
        return method.method_id
    except Exception as e:
        db.rollback()
        logger.error("Method 入图失败: %s", e)
        raise
    finally:
        db.close()


def add_dataset(name: str, domain: str = "", task: str = "",
                description: str = "", year: int = None) -> str:
    """创建或查找已有 Dataset（按 name 去重）"""
    db = SessionLocal()
    try:
        existing = db.query(KgDataset).filter(KgDataset.name == name).first()
        if existing:
            # 补充信息
            if domain and not existing.domain:
                existing.domain = domain
            if task and not existing.task:
                existing.task = task
            if year and not existing.year:
                existing.year = year
            db.commit()
            return existing.dataset_id

        dataset = KgDataset(name=name, domain=domain, task=task,
                            description=description, year=year)
        db.add(dataset)
        db.commit()
        _mark_graph_dirty()
        return dataset.dataset_id
    except Exception as e:
        db.rollback()
        logger.error("Dataset 入图失败: %s", e)
        raise
    finally:
        db.close()


def add_task(name: str, description: str = "", parent_task_id: str = None,
             level: int = 1) -> str:
    """创建或查找已有 ResearchTask"""
    db = SessionLocal()
    try:
        existing = db.query(KgTask).filter(KgTask.name == name).first()
        if existing:
            return existing.task_id

        task = KgTask(name=name, description=description,
                      parent_task_id=parent_task_id, level=level)
        db.add(task)
        db.commit()
        _mark_graph_dirty()
        return task.task_id
    except Exception as e:
        db.rollback()
        logger.error("Task 入图失败: %s", e)
        raise
    finally:
        db.close()


def add_metric(name: str, full_name: str = "", description: str = "",
               direction: str = "", unit: str = "") -> str:
    """创建或查找已有 Metric"""
    db = SessionLocal()
    try:
        existing = db.query(KgMetric).filter(KgMetric.name == name).first()
        if existing:
            return existing.metric_id

        from knowledge_graph.models import MetricDirection
        d = None
        if direction:
            try:
                d = MetricDirection(direction)
            except ValueError:
                pass
        metric = KgMetric(name=name, full_name=full_name, description=description,
                          direction=d, unit=unit)
        db.add(metric)
        db.commit()
        _mark_graph_dirty()
        return metric.metric_id
    except Exception as e:
        db.rollback()
        logger.error("Metric 入图失败: %s", e)
        raise
    finally:
        db.close()


def add_venue(name: str, abbreviation: str = "", venue_type: str = "",
              rank: str = "", publisher: str = "") -> str:
    """创建或查找已有 Venue"""
    db = SessionLocal()
    try:
        existing = db.query(KgVenue).filter(KgVenue.name == name).first()
        if existing:
            return existing.venue_id

        from knowledge_graph.models import VenueType
        vt = None
        if venue_type:
            try:
                vt = VenueType(venue_type)
            except ValueError:
                pass
        venue = KgVenue(name=name, abbreviation=abbreviation, type=vt,
                        rank=rank, publisher=publisher)
        db.add(venue)
        db.commit()
        _mark_graph_dirty()
        return venue.venue_id
    except Exception as e:
        db.rollback()
        logger.error("Venue 入图失败: %s", e)
        raise
    finally:
        db.close()


# ============================================================
# 关系 CRUD
# ============================================================

def link_paper_author(paper_id: str, author_id: str, author_order: int = 1,
                      is_corresponding: bool = False):
    """建立 WRITTEN_BY 关系"""
    db = SessionLocal()
    try:
        existing = db.query(KgPaperAuthor).filter(
            KgPaperAuthor.paper_id == paper_id,
            KgPaperAuthor.author_id == author_id,
        ).first()
        if existing:
            return
        db.add(KgPaperAuthor(paper_id=paper_id, author_id=author_id,
                             author_order=author_order, is_corresponding=is_corresponding))
        db.commit()
        _mark_graph_dirty()
    except Exception as e:
        db.rollback()
        logger.error("link_paper_author 失败: %s", e)
    finally:
        db.close()


def link_paper_method(paper_id: str, method_id: str, role: str = "",
                      variant: str = "", context: str = "",
                      performance_contribution: str = ""):
    """建立 USES_METHOD 关系"""
    db = SessionLocal()
    try:
        existing = db.query(KgPaperUsesMethod).filter(
            KgPaperUsesMethod.paper_id == paper_id,
            KgPaperUsesMethod.method_id == method_id,
        ).first()
        if existing:
            return
        from knowledge_graph.models import PerformanceContribution
        pc = None
        if performance_contribution:
            try:
                pc = PerformanceContribution(performance_contribution)
            except ValueError:
                pass
        db.add(KgPaperUsesMethod(paper_id=paper_id, method_id=method_id,
                                 role=role, variant=variant, context=context,
                                 performance_contribution=pc))
        db.commit()
        _mark_graph_dirty()
    except Exception as e:
        db.rollback()
        logger.error("link_paper_method 失败: %s", e)
    finally:
        db.close()


def link_paper_dataset_eval(paper_id: str, dataset_id: str, task: str = "",
                            split: str = "eval", metrics: Dict = None,
                            protocol: str = ""):
    """建立 EVALUATES_ON 关系"""
    db = SessionLocal()
    try:
        existing = db.query(KgPaperEvaluatesOn).filter(
            KgPaperEvaluatesOn.paper_id == paper_id,
            KgPaperEvaluatesOn.dataset_id == dataset_id,
        ).first()
        if existing:
            # 合并 metrics
            if metrics:
                cur = existing.metrics or {}
                if isinstance(cur, str):
                    cur = json.loads(cur)
                cur.update(metrics)
                existing.metrics = cur
                db.commit()
            return
        db.add(KgPaperEvaluatesOn(paper_id=paper_id, dataset_id=dataset_id,
                                  task=task, split=split, metrics=metrics or {},
                                  protocol=protocol))
        db.commit()
        _mark_graph_dirty()
    except Exception as e:
        db.rollback()
        logger.error("link_paper_dataset_eval 失败: %s", e)
    finally:
        db.close()


def link_paper_dataset_train(paper_id: str, dataset_id: str, split: str = "train+dev"):
    """建立 TRAINS_ON 关系"""
    db = SessionLocal()
    try:
        existing = db.query(KgPaperTrainsOn).filter(
            KgPaperTrainsOn.paper_id == paper_id,
            KgPaperTrainsOn.dataset_id == dataset_id,
        ).first()
        if existing:
            return
        db.add(KgPaperTrainsOn(paper_id=paper_id, dataset_id=dataset_id, split=split))
        db.commit()
        _mark_graph_dirty()
    except Exception as e:
        db.rollback()
        logger.error("link_paper_dataset_train 失败: %s", e)
    finally:
        db.close()


def link_paper_task(paper_id: str, task_id: str, is_primary: bool = True):
    """建立 BELONGS_TO 关系"""
    db = SessionLocal()
    try:
        existing = db.query(KgPaperBelongsToTask).filter(
            KgPaperBelongsToTask.paper_id == paper_id,
            KgPaperBelongsToTask.task_id == task_id,
        ).first()
        if existing:
            return
        db.add(KgPaperBelongsToTask(paper_id=paper_id, task_id=task_id,
                                    is_primary=is_primary))
        db.commit()
        _mark_graph_dirty()
    except Exception as e:
        db.rollback()
        logger.error("link_paper_task 失败: %s", e)
    finally:
        db.close()


def link_paper_venue(paper_id: str, venue_id: str, date: str = ""):
    """建立 PUBLISHED_IN 关系"""
    db = SessionLocal()
    try:
        existing = db.query(KgPaperPublishedIn).filter(
            KgPaperPublishedIn.paper_id == paper_id,
            KgPaperPublishedIn.venue_id == venue_id,
        ).first()
        if existing:
            return
        db.add(KgPaperPublishedIn(paper_id=paper_id, venue_id=venue_id, date=date))
        db.commit()
        _mark_graph_dirty()
    except Exception as e:
        db.rollback()
        logger.error("link_paper_venue 失败: %s", e)
    finally:
        db.close()


def link_paper_metric(paper_id: str, metric_id: str, value: float = None,
                      dataset_id: str = None, condition: str = "", notes: str = ""):
    """建立 REPORTS_METRIC 关系"""
    db = SessionLocal()
    try:
        db.add(KgPaperReportsMetric(paper_id=paper_id, metric_id=metric_id,
                                    value=value, dataset_id=dataset_id,
                                    condition=condition, notes=notes))
        db.commit()
        _mark_graph_dirty()
    except Exception as e:
        db.rollback()
        logger.error("link_paper_metric 失败: %s", e)
    finally:
        db.close()


# ============================================================
# 图查询辅助
# ============================================================

def get_node(entity_type: str, entity_id: str) -> Optional[Dict]:
    """查找单个实体"""
    type_to_model = {
        "paper": KgPaper, "author": KgAuthor, "dataset": KgDataset,
        "method": KgMethod, "metric": KgMetric, "task": KgTask, "venue": KgVenue,
    }
    model = type_to_model.get(entity_type)
    if not model:
        return None

    db = SessionLocal()
    try:
        pk_map = {
            "paper": KgPaper.paper_id, "author": KgAuthor.author_id,
            "dataset": KgDataset.dataset_id, "method": KgMethod.method_id,
            "metric": KgMetric.metric_id, "task": KgTask.task_id,
            "venue": KgVenue.venue_id,
        }
        pk_col = pk_map.get(entity_type)
        obj = db.query(model).filter(pk_col == entity_id).first()
        if obj is None:
            return None
        return _model_to_dict(obj, entity_type)
    finally:
        db.close()


def search_nodes(entity_type: str, name_pattern: str, limit: int = 20) -> List[Dict]:
    """按名称模糊搜索实体"""
    type_to_model = {
        "paper": (KgPaper, KgPaper.title, "paper_id"),
        "author": (KgAuthor, KgAuthor.name, "author_id"),
        "dataset": (KgDataset, KgDataset.name, "dataset_id"),
        "method": (KgMethod, KgMethod.name, "method_id"),
        "metric": (KgMetric, KgMetric.name, "metric_id"),
        "task": (KgTask, KgTask.name, "task_id"),
        "venue": (KgVenue, KgVenue.name, "venue_id"),
    }
    entry = type_to_model.get(entity_type)
    if not entry:
        return []

    model, col, id_col = entry
    db = SessionLocal()
    try:
        results = db.query(model).filter(col.contains(name_pattern)).limit(limit).all()
        return [_model_to_dict(r, entity_type) for r in results]
    finally:
        db.close()


def get_graph_stats() -> Dict:
    """获取图谱统计信息"""
    db = SessionLocal()
    try:
        return {
            "papers": db.query(KgPaper).count(),
            "authors": db.query(KgAuthor).count(),
            "datasets": db.query(KgDataset).count(),
            "methods": {
                "feature_extractor": db.query(KgMethod).filter(KgMethod.category == "feature_extractor").count(),
                "network_architecture": db.query(KgMethod).filter(KgMethod.category == "network_architecture").count(),
                "loss_function": db.query(KgMethod).filter(KgMethod.category == "loss_function").count(),
            },
            "metrics": db.query(KgMetric).count(),
            "tasks": db.query(KgTask).count(),
            "venues": db.query(KgVenue).count(),
            "relationships": {
                "paper_author": db.query(KgPaperAuthor).count(),
                "paper_cites": db.query(KgPaperCites).count(),
                "paper_uses_method": db.query(KgPaperUsesMethod).count(),
                "paper_evaluates_on": db.query(KgPaperEvaluatesOn).count(),
                "paper_trains_on": db.query(KgPaperTrainsOn).count(),
                "paper_belongs_to_task": db.query(KgPaperBelongsToTask).count(),
                "method_improves_method": db.query(KgMethodImprovesMethod).count(),
                "paper_published_in": db.query(KgPaperPublishedIn).count(),
                "paper_reports_metric": db.query(KgPaperReportsMetric).count(),
            },
        }
    finally:
        db.close()


def clear_graph():
    """清空所有知识图谱数据（实体 + 关系）"""
    from knowledge_graph.models import (
        KgPaper, KgAuthor, KgDataset, KgMethod, KgMetric, KgTask, KgVenue,
        KgPaperAuthor, KgPaperCites, KgPaperUsesMethod,
        KgPaperEvaluatesOn, KgPaperTrainsOn, KgPaperBelongsToTask,
        KgMethodImprovesMethod, KgPaperPublishedIn, KgPaperReportsMetric,
    )
    global _nx_graph, _graph_dirty
    db = SessionLocal()
    try:
        # 先删关系表，再删实体表（避免 FK 约束问题）
        for rel_model in [KgPaperAuthor, KgPaperCites, KgPaperUsesMethod,
                          KgPaperEvaluatesOn, KgPaperTrainsOn, KgPaperBelongsToTask,
                          KgMethodImprovesMethod, KgPaperPublishedIn, KgPaperReportsMetric]:
            db.query(rel_model).delete()
        for entity_model in [KgPaper, KgAuthor, KgDataset, KgMethod, KgMetric, KgTask, KgVenue]:
            db.query(entity_model).delete()
        db.commit()
        logger.info("知识图谱已清空")
    except Exception as e:
        db.rollback()
        logger.error("清空知识图谱失败: %s", e)
        raise
    finally:
        db.close()
    _nx_graph = None
    _graph_dirty = True


def _model_to_dict(obj, entity_type: str) -> Dict:
    """ORM 对象转字典"""
    if entity_type == "paper":
        return {
            "paper_id": obj.paper_id, "title": obj.title,
            "abstract": obj.abstract, "abstract_cn": obj.abstract_cn,
            "keywords": obj.keywords, "authors": obj.authors or [],
            "method_name": obj.method_name or "",
            "method_summary": obj.method_summary or "",
            "year": obj.year,
            "doi": obj.doi, "arxiv_id": obj.arxiv_id,
            "citation_count": obj.citation_count,
            "venue_name": obj.venue_name, "source": obj.source.value if obj.source else "",
        }
    elif entity_type == "author":
        return {"author_id": obj.author_id, "name": obj.name,
                "affiliation": obj.affiliation, "orcid": obj.orcid}
    elif entity_type == "dataset":
        return {"dataset_id": obj.dataset_id, "name": obj.name,
                "domain": obj.domain, "task": obj.task, "year": obj.year}
    elif entity_type == "method":
        return {"method_id": obj.method_id, "name": obj.name,
                "aliases": obj.aliases,
                "category": obj.category or "",
                "description": obj.description}
    elif entity_type == "task":
        return {"task_id": obj.task_id, "name": obj.name,
                "description": obj.description,
                "parent_task_id": obj.parent_task_id, "level": obj.level}
    elif entity_type == "metric":
        return {"metric_id": obj.metric_id, "name": obj.name,
                "full_name": obj.full_name,
                "direction": obj.direction.value if obj.direction else ""}
    elif entity_type == "venue":
        return {"venue_id": obj.venue_id, "name": obj.name,
                "abbreviation": obj.abbreviation,
                "type": obj.type.value if obj.type else ""}
    return {}


# ============================================================
# Excel 导出
# ============================================================

def export_papers_to_excel() -> Tuple[Optional[bytes], str]:
    """
    将知识库所有论文信息导出为 Excel 文件。

    返回:
        (excel_bytes, filename) 或 (None, error_message)
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    db = SessionLocal()
    try:
        papers = db.query(KgPaper).order_by(KgPaper.year.desc(), KgPaper.created_at.desc()).all()
        if not papers:
            return None, "知识库中暂无论文"

        wb = Workbook()
        ws = wb.active
        ws.title = "论文信息"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # 表头
        headers = [
            "序号", "论文标题", "关键词", "发表年份", "发表地",
            "提出方法名称", "提出方法概述",
            "特征提取器", "网络框架", "损失函数",
            "实验数据集",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 数据行
        for i, p in enumerate(papers):
            row = i + 2
            paper_id = p.paper_id

            # 关联数据查询
            methods = (
                db.query(KgMethod, KgPaperUsesMethod)
                .join(KgPaperUsesMethod, KgPaperUsesMethod.method_id == KgMethod.method_id)
                .filter(KgPaperUsesMethod.paper_id == paper_id)
                .all()
            )
            datasets_eval = (
                db.query(KgDataset, KgPaperEvaluatesOn)
                .join(KgPaperEvaluatesOn, KgPaperEvaluatesOn.dataset_id == KgDataset.dataset_id)
                .filter(KgPaperEvaluatesOn.paper_id == paper_id)
                .all()
            )
            datasets_train = (
                db.query(KgDataset, KgPaperTrainsOn)
                .join(KgPaperTrainsOn, KgPaperTrainsOn.dataset_id == KgDataset.dataset_id)
                .filter(KgPaperTrainsOn.paper_id == paper_id)
                .all()
            )

            # 分类方法
            fes = [m for m, r in methods if r.role == "feature_extractor" or m.category == "feature_extractor"]
            nets = [m for m, r in methods if r.role == "network_architecture" or m.category == "network_architecture"]
            losses = [m for m, r in methods if r.role == "loss_function" or m.category == "loss_function"]

            # 组装每列数据
            keywords = p.keywords or []

            proposed_names = p.method_name or ""
            proposed_summaries = (p.method_summary or "")[:500]

            fe_str = "; ".join(f"{m.name}" + (f"({r.variant})" if (r := next((rr for mm, rr in methods if mm.method_id == m.method_id), None)) and r.variant else "") for m in fes)
            net_str = "; ".join(m.name for m in nets)
            loss_str = "; ".join(m.name for m in losses)

            ds_parts = []
            for ds, rel in datasets_eval:
                ds_parts.append(f"{ds.name}(评测, {rel.task or ''})")
            for ds, rel in datasets_train:
                ds_parts.append(f"{ds.name}(训练, {rel.split or ''})")
            ds_str = "; ".join(ds_parts)

            row_data = [
                i + 1,
                p.title,
                ", ".join(keywords) if keywords else "",
                p.year or "",
                p.venue_name or "",
                proposed_names,
                proposed_summaries[:500] if proposed_summaries else "",
                fe_str,
                net_str,
                loss_str,
                ds_str,
            ]

            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # 调整列宽
        col_widths = [5, 40, 20, 8, 18, 20, 40, 25, 20, 20, 30]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

        # 冻结首行
        ws.freeze_panes = "A2"

        # 写入 bytes buffer
        buffer = _io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"知识库论文信息_{len(papers)}篇.xlsx"
        logger.info("Excel 导出完成: %d 篇论文", len(papers))
        return buffer.getvalue(), filename

    except Exception as e:
        logger.error("Excel 导出失败: %s", e, exc_info=True)
        return None, f"导出失败: {e}"
    finally:
        db.close()


# ============================================================
# 知识库管理：论文查询与删除
# ============================================================

def get_all_papers(limit: int = 200) -> List[Dict]:
    """
    获取知识库中所有论文的基本信息。
    返回按年份倒序排列的论文列表。
    """
    db = SessionLocal()
    try:
        papers = (
            db.query(KgPaper)
            .order_by(KgPaper.year.is_(None), KgPaper.year.desc(), KgPaper.created_at.desc())
            .limit(limit)
            .all()
        )
        result = []
        for p in papers:
            # 统计关联实体数
            method_count = db.query(KgPaperUsesMethod).filter(
                KgPaperUsesMethod.paper_id == p.paper_id
            ).count()
            dataset_count = db.query(KgPaperEvaluatesOn).filter(
                KgPaperEvaluatesOn.paper_id == p.paper_id
            ).count()
            task_count = db.query(KgPaperBelongsToTask).filter(
                KgPaperBelongsToTask.paper_id == p.paper_id
            ).count()
            author_count = db.query(KgPaperAuthor).filter(
                KgPaperAuthor.paper_id == p.paper_id
            ).count()

            result.append({
                "paper_id": p.paper_id,
                "title": p.title,
                "year": p.year,
                "venue_name": p.venue_name,
                "keywords": p.keywords or [],
                "authors": p.authors or [],
                "method_name": p.method_name or "",
                "abstract": (p.abstract or "")[:500],
                "source": p.source.value if p.source else "upload",
                "document_id": p.document_id,
                "created_at": p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "",
                "method_count": method_count,
                "dataset_count": dataset_count,
                "task_count": task_count,
                "author_count": author_count,
            })
        return result
    finally:
        db.close()


def get_paper_detail(paper_id: str) -> Optional[Dict]:
    """
    获取单篇论文的完整信息，包括关联的所有实体和关系。
    """
    db = SessionLocal()
    try:
        paper = db.query(KgPaper).filter(KgPaper.paper_id == paper_id).first()
        if not paper:
            return None

        # 基本信息
        detail = {
            "paper_id": paper.paper_id,
            "title": paper.title,
            "year": paper.year,
            "venue_name": paper.venue_name,
            "keywords": paper.keywords or [],
            "authors": paper.authors or [],
            "method_name": paper.method_name or "",
            "method_summary": paper.method_summary or "",
            "abstract": paper.abstract or "",
            "abstract_cn": paper.abstract_cn or "",
            "doi": paper.doi,
            "arxiv_id": paper.arxiv_id,
            "citation_count": paper.citation_count or 0,
            "source": paper.source.value if paper.source else "upload",
            "document_id": paper.document_id,
            "created_at": paper.created_at.strftime("%Y-%m-%d %H:%M") if paper.created_at else "",
        }

        # 方法（含 feature_extractor / network_architecture / loss_function / proposed）
        methods = (
            db.query(KgMethod, KgPaperUsesMethod)
            .join(KgPaperUsesMethod, KgPaperUsesMethod.method_id == KgMethod.method_id)
            .filter(KgPaperUsesMethod.paper_id == paper_id)
            .all()
        )
        detail["methods"] = [
            {
                "method_id": m.method_id, "name": m.name,
                "category": m.category or rel.role or "",
                "description": (m.description or "")[:300],
                "role": rel.role or "",
                "variant": rel.variant or "",
            }
            for m, rel in methods
        ]

        # 数据集（评测）
        datasets_eval = (
            db.query(KgDataset, KgPaperEvaluatesOn)
            .join(KgPaperEvaluatesOn, KgPaperEvaluatesOn.dataset_id == KgDataset.dataset_id)
            .filter(KgPaperEvaluatesOn.paper_id == paper_id)
            .all()
        )
        datasets_train = (
            db.query(KgDataset, KgPaperTrainsOn)
            .join(KgPaperTrainsOn, KgPaperTrainsOn.dataset_id == KgDataset.dataset_id)
            .filter(KgPaperTrainsOn.paper_id == paper_id)
            .all()
        )
        detail["datasets"] = []
        for ds, rel in datasets_eval:
            detail["datasets"].append({
                "dataset_id": ds.dataset_id, "name": ds.name,
                "description": (ds.description or "")[:300],
                "role": "eval",
                "task": rel.task or "",
                "split": rel.split or "",
            })
        for ds, rel in datasets_train:
            detail["datasets"].append({
                "dataset_id": ds.dataset_id, "name": ds.name,
                "description": (ds.description or "")[:300],
                "role": "train",
                "task": "",
                "split": rel.split or "",
            })

        # 任务
        tasks = (
            db.query(KgTask, KgPaperBelongsToTask)
            .join(KgPaperBelongsToTask, KgPaperBelongsToTask.task_id == KgTask.task_id)
            .filter(KgPaperBelongsToTask.paper_id == paper_id)
            .all()
        )
        detail["tasks"] = [
            {
                "task_id": t.task_id, "name": t.name,
                "description": (t.description or "")[:200],
                "is_primary": rel.is_primary,
            }
            for t, rel in tasks
        ]

        # 指标（从 KgPaperReportsMetric 和 EvaluatesOn.metrics JSON 两处获取）
        detail["metrics"] = []
        # 方式1: KgPaperReportsMetric 表
        metrics_rows = (
            db.query(KgMetric, KgPaperReportsMetric)
            .join(KgPaperReportsMetric, KgPaperReportsMetric.metric_id == KgMetric.metric_id)
            .filter(KgPaperReportsMetric.paper_id == paper_id)
            .all()
        )
        for m, rel in metrics_rows:
            detail["metrics"].append({
                "metric_id": m.metric_id, "name": m.name,
                "value": rel.value,
                "dataset_id": rel.dataset_id,
                "condition": rel.condition or "",
                "notes": (rel.notes or "")[:200],
            })
        # 方式2: EvaluatesOn.metrics JSON（新论文的实验结果存在这里）
        eval_rels = (
            db.query(KgPaperEvaluatesOn, KgDataset)
            .join(KgDataset, KgDataset.dataset_id == KgPaperEvaluatesOn.dataset_id)
            .filter(KgPaperEvaluatesOn.paper_id == paper_id)
            .all()
        )
        for er, ds in eval_rels:
            metrics_json = er.metrics or {}
            if isinstance(metrics_json, str):
                try:
                    metrics_json = json.loads(metrics_json)
                except Exception:
                    metrics_json = {}
            for condition, table in metrics_json.items():
                if isinstance(table, list):
                    for row in table:
                        method_name = row.get("method_name", "")
                        for m_name, m_val in row.get("metrics", {}).items():
                            detail["metrics"].append({
                                "metric_id": "", "name": m_name,
                                "value": m_val,
                                "dataset_id": er.dataset_id,
                                "condition": f"{condition} | {method_name}",
                                "notes": f"{'baseline' if row.get('is_baseline') else 'proposed'}",
                            })

        # 引用关系
        cites_count = db.query(KgPaperCites).filter(
            KgPaperCites.citing_paper_id == paper_id
        ).count()
        cited_by_count = db.query(KgPaperCites).filter(
            KgPaperCites.cited_paper_id == paper_id
        ).count()
        detail["cites_count"] = cites_count
        detail["cited_by_count"] = cited_by_count

        return detail
    finally:
        db.close()


def delete_paper(paper_id: str) -> Tuple[bool, str]:
    """
    删除一篇论文及其所有关联关系。
    同时清理关联的 documents/chunks 表记录、上传文件和向量数据。

    返回:
        (success, message)
    """
    import os as _os
    from config import UPLOAD_DIR
    from database.models import Document, Chunk
    from knowledge_base.vector_store import delete_document_chunks

    db = SessionLocal()
    try:
        paper = db.query(KgPaper).filter(KgPaper.paper_id == paper_id).first()
        if not paper:
            return False, f"论文不存在: {paper_id}"

        title = paper.title[:80]
        document_id = paper.document_id

        # 1. 收集需要清理的文件路径（在删 DB 记录前获取）
        file_paths_to_remove = []
        if document_id:
            doc = db.query(Document).filter(Document.id == document_id).first()
            if doc:
                if doc.file_path and _os.path.exists(doc.file_path):
                    file_paths_to_remove.append(doc.file_path)
                # 同时收集关联的 MD 文件
                md_path = _os.path.join(UPLOAD_DIR, f"{doc.md5_hash}.md")
                if _os.path.exists(md_path):
                    file_paths_to_remove.append(md_path)

        # 2. 删除所有关系（手动删除更安全，避免 CASCADE 的隐式行为）
        # CITES 表有两个 paper_id 列
        db.query(KgPaperCites).filter(
            (KgPaperCites.citing_paper_id == paper_id) |
            (KgPaperCites.cited_paper_id == paper_id)
        ).delete()
        db.query(KgPaperAuthor).filter(KgPaperAuthor.paper_id == paper_id).delete()
        db.query(KgPaperUsesMethod).filter(KgPaperUsesMethod.paper_id == paper_id).delete()
        db.query(KgPaperEvaluatesOn).filter(KgPaperEvaluatesOn.paper_id == paper_id).delete()
        db.query(KgPaperTrainsOn).filter(KgPaperTrainsOn.paper_id == paper_id).delete()
        db.query(KgPaperBelongsToTask).filter(KgPaperBelongsToTask.paper_id == paper_id).delete()
        db.query(KgPaperReportsMetric).filter(KgPaperReportsMetric.paper_id == paper_id).delete()
        db.query(KgPaperPublishedIn).filter(KgPaperPublishedIn.paper_id == paper_id).delete()
        db.query(KgMethodImprovesMethod).filter(
            KgMethodImprovesMethod.paper_id == paper_id
        ).delete()

        # 3. 删除 Paper 实体
        db.delete(paper)

        # 4. 清理关联的 documents 表和 chunks 表
        if document_id:
            db.query(Chunk).filter(Chunk.document_id == document_id).delete()
            db.query(Document).filter(Document.id == document_id).delete()
            # 从向量库删除
            try:
                delete_document_chunks(document_id)
            except Exception as e:
                logger.warning("向量库删除失败: %s", e)

        db.commit()
        _mark_graph_dirty()

        # 5. 删除磁盘文件
        for fpath in file_paths_to_remove:
            try:
                _os.remove(fpath)
                logger.info("已删除文件: %s", fpath)
            except OSError as e:
                logger.warning("删除文件失败: %s — %s", fpath, e)

        logger.info("论文已删除: '%s' (paper_id=%s, doc_id=%s)", title, paper_id[:12], document_id)
        return True, f"已删除论文「{title}」及其关联数据"

    except Exception as e:
        db.rollback()
        logger.error("删除论文失败: %s", e, exc_info=True)
        return False, f"删除失败: {e}"
    finally:
        db.close()
